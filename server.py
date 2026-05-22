from __future__ import annotations

import base64
import hashlib
import json
import os
import re
import secrets
import ssl
import sys
import urllib.parse
import urllib.request
import uuid
from dataclasses import dataclass
from datetime import datetime, timedelta, timezone
from email.message import EmailMessage
from pathlib import Path
from typing import Any

from flask import Flask, jsonify, redirect, request, send_from_directory, session
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent
STATIC_DIR = ROOT / "static"
DATA_DIR = ROOT / "data"
UPLOADS_DIR = ROOT / "uploads"
DEFAULT_WORKBOOK = ROOT / "report_centric_investor_demo_dataset.xlsx"
CURRENT_WORKBOOK_PATH: Path | None = None
LAST_UPLOADED_AT = ""
REQUIRED_SHEETS = ["Companies", "Investors", "Report_Recipients"]

REPORT_ORDER = ["monthly", "quarterly", "annual"]
REPORT_LABELS = {
    "monthly": {"zh": "月報", "en": "Monthly Report"},
    "quarterly": {"zh": "季報", "en": "Quarterly Report"},
    "annual": {"zh": "年報", "en": "Annual Report"},
}

SIMPLE_REPORTS = {
    "monthly": {"zh": "月報", "en": "Monthly Report", "path": UPLOADS_DIR / "reports" / "monthly.pdf"},
    "quarterly": {"zh": "季報", "en": "Quarterly Report", "path": UPLOADS_DIR / "reports" / "quarterly.pdf"},
    "annual": {"zh": "年報", "en": "Annual Report", "path": UPLOADS_DIR / "reports" / "annual.pdf"},
}

DEFAULT_TEMPLATES = {
    "zh": {
        "name": "中文預設模板",
        "language": "zh",
        "subject": "{{fund_name}} {{report_period}} 投資人報告",
        "body": "親愛的 {{nickname}} 您好，\n\n隨信附上 {{fund_name}} 之 {{report_period}} 投資人報告，敬請參閱。\n\n本次附件包含：\n{{attachment_list}}\n\n若有任何問題，請隨時與我們聯繫。\n\n敬祝 順心",
    },
    "en": {
        "name": "英文預設模板",
        "language": "en",
        "subject": "{{fund_name}} {{report_period}} Investor Reports",
        "body": "Dear {{nickname}},\n\nPlease find attached the {{report_period}} investor report(s) for {{fund_name}}.\n\nThe attached files include:\n{{attachment_list}}\n\nPlease let us know if you have any questions.\n\nBest regards,",
    },
}


def load_env_file() -> None:
    env_path = ROOT / ".env"
    if not env_path.exists():
        return
    for line in env_path.read_text(encoding="utf-8").splitlines():
        stripped = line.strip()
        if not stripped or stripped.startswith("#") or "=" not in stripped:
            continue
        key, value = stripped.split("=", 1)
        os.environ.setdefault(key.strip(), value.strip().strip('"').strip("'"))


def load_runtime_workbook_metadata() -> None:
    global CURRENT_WORKBOOK_PATH, LAST_UPLOADED_AT

    meta_path = DATA_DIR / "current_database.meta.json"
    workbook_file = DATA_DIR / "current_database.xlsx"
    if not meta_path.exists() or not workbook_file.exists():
        return
    try:
        metadata = json.loads(meta_path.read_text(encoding="utf-8"))
        CURRENT_WORKBOOK_PATH = workbook_file
        LAST_UPLOADED_AT = str(metadata.get("uploaded_at") or "")
    except Exception:
        CURRENT_WORKBOOK_PATH = workbook_file


@dataclass
class UploadedFile:
    field_name: str
    filename: str
    content_type: str
    content: bytes


@dataclass
class SelectedReport:
    company_id: str
    company_name: str
    report_type: str
    file: UploadedFile


SAMPLE_INVESTORS = [
    {
        "investor_id": "INV-001",
        "investor_name": "Demo Capital Partners",
        "email": "demo.capital@example.com",
        "cc": "ops@example.com",
        "language": "en",
        "status": "active",
    },
    {
        "investor_id": "INV-002",
        "investor_name": "宏遠投資有限公司",
        "email": "investor.zh@example.com",
        "cc": "",
        "language": "zh",
        "status": "active",
    },
    {
        "investor_id": "INV-003",
        "investor_name": "Inactive Demo Investor",
        "email": "inactive@example.com",
        "cc": "",
        "language": "en",
        "status": "inactive",
    },
]

SAMPLE_COMPANIES = [
    {"company_id": "C001", "company_name": "Company A", "short_name": "A", "status": "active"},
    {"company_id": "C002", "company_name": "Company B", "short_name": "B", "status": "active"},
    {"company_id": "C003", "company_name": "Company C", "short_name": "C", "status": "active"},
]

SAMPLE_REPORT_RECIPIENTS = [
    {"company_id": "C001", "report_type": "monthly", "investor_id": "INV-001", "active": True},
    {"company_id": "C001", "report_type": "monthly", "investor_id": "INV-002", "active": True},
    {"company_id": "C002", "report_type": "monthly", "investor_id": "INV-002", "active": True},
    {"company_id": "C002", "report_type": "monthly", "investor_id": "INV-003", "active": True},
    {"company_id": "C003", "report_type": "annual", "investor_id": "INV-001", "active": True},
]


def workbook_path() -> Path:
    if CURRENT_WORKBOOK_PATH is not None:
        return CURRENT_WORKBOOK_PATH
    configured = os.getenv("INVESTOR_WORKBOOK_PATH")
    if configured:
        path = Path(configured)
        return path if path.is_absolute() else ROOT / path
    return DEFAULT_WORKBOOK


def workbook_state() -> dict[str, Any]:
    path = workbook_path()
    if not path.exists():
        return {
            "path": str(path),
            "usingFallback": True,
            "warning": "目前使用的是系統內建示範資料，尚未成功讀取 Excel 資料庫。",
            "error": f"Workbook not found: {path}",
            "sheetNames": [],
            "missingRequiredSheets": REQUIRED_SHEETS,
        }
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet_names = list(workbook.sheetnames)
        workbook.close()
        missing = [sheet for sheet in REQUIRED_SHEETS if sheet not in sheet_names]
        return {
            "path": str(path),
            "usingFallback": False,
            "warning": f"資料庫 Excel 缺少必要工作表：{' / '.join(missing)}" if missing else "",
            "error": "",
            "sheetNames": sheet_names,
            "missingRequiredSheets": missing,
        }
    except Exception as error:
        return {
            "path": str(path),
            "usingFallback": True,
            "warning": "目前使用的是系統內建示範資料，尚未成功讀取 Excel 資料庫。",
            "error": str(error),
            "sheetNames": [],
            "missingRequiredSheets": REQUIRED_SHEETS,
        }


def normalize_language(value: Any) -> str:
    text = str(value or "").strip().lower()
    return "en" if text in {"en", "english"} else "zh"


def normalize_status(value: Any) -> str:
    text = str(value or "").strip().lower()
    if text in {"inactive", "停用", "不啟用", "disabled", "disable"}:
        return "inactive"
    return "active"


def normalize_recipient_row(row: dict[str, Any], index: int) -> dict[str, str]:
    recipient_id = str(first_value(row, "recipient_id", "Recipient_ID", "investor_id", "Investor_ID")).strip()
    if not recipient_id:
        recipient_id = f"REC{index:03d}"
    name = str(first_value(row, "name", "recipient_name", "investor_name", "Name", "Recipient_Name", "Investor_Name")).strip()
    nickname = str(first_value(row, "nickname", "display_name", "Nickname", "Display_Name")).strip()
    return {
        "recipient_id": recipient_id,
        "name": name,
        "nickname": nickname,
        "email": str(first_value(row, "email", "Email", "mail", "Mail")).strip(),
        "cc": str(first_value(row, "cc", "CC")).strip(),
        "language": normalize_language(first_value(row, "language", "lang", "Language", "Lang")),
        "status": normalize_status(first_value(row, "status", "Status")),
        "notes": str(first_value(row, "notes", "remark", "Notes", "Remark")).strip(),
    }


def parse_recipients_workbook(path: Path) -> list[dict[str, str]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook["Recipients"] if "Recipients" in workbook.sheetnames else workbook[workbook.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        workbook.close()
        return []
    headers = [str(cell or "").strip() for cell in rows[0]]
    recipients = []
    for index, row in enumerate(rows[1:], start=1):
        record = {headers[column_index]: row[column_index] if column_index < len(row) else "" for column_index in range(len(headers))}
        recipient = normalize_recipient_row(record, index)
        if recipient["name"] or recipient["email"]:
            recipients.append(recipient)
    workbook.close()
    return recipients


def recipients_summary(recipients: list[dict[str, str]]) -> dict[str, int]:
    return {
        "total": len(recipients),
        "active": len([item for item in recipients if item["status"] == "active"]),
        "inactive": len([item for item in recipients if item["status"] != "active"]),
        "missing_email": len([item for item in recipients if not item["email"]]),
    }


def load_investors() -> list[dict[str, str]]:
    path = workbook_path()
    state = workbook_state()
    if state["usingFallback"]:
        return SAMPLE_INVESTORS

    workbook = load_workbook(path, read_only=True, data_only=True)
    if "Investors" not in workbook.sheetnames:
        workbook.close()
        return []

    sheet = workbook["Investors"]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(cell or "").strip() for cell in rows[0]]
    investors = []
    for row in rows[1:]:
        record = {headers[index]: row[index] if index < len(row) else "" for index in range(len(headers))}
        investor = {
            "investor_id": str(first_value(record, "investor_id", "Investor_ID")).strip(),
            "investor_name": str(first_value(record, "investor_name", "Investor_Name", "name")).strip(),
            "email": str(first_value(record, "email", "Email")).strip(),
            "cc": str(first_value(record, "cc", "CC")).strip(),
            "language": normalize_language(first_value(record, "language", "Language")),
            "status": str(first_value(record, "status", "Status") or "active").strip().lower(),
        }
        if investor["investor_id"] and investor["investor_name"]:
            investors.append(investor)
    workbook.close()
    return investors


def active_investors() -> list[dict[str, str]]:
    return [investor for investor in load_investors() if investor["status"] == "active"]


def sheet_rows(sheet_name: str) -> list[dict[str, Any]]:
    path = workbook_path()
    state = workbook_state()
    if state["usingFallback"]:
        return []
    workbook = load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        workbook.close()
        return []
    rows = list(workbook[sheet_name].iter_rows(values_only=True))
    if not rows:
        workbook.close()
        return []
    headers = [str(cell or "").strip() for cell in rows[0]]
    output = [
        {headers[index]: row[index] if index < len(row) else "" for index in range(len(headers))}
        for row in rows[1:]
    ]
    workbook.close()
    return output


def truthy(value: Any) -> bool:
    text = str(value).strip().lower()
    return text in {"1", "true", "yes", "y", "active", "是"}


def normalize_keyed_row(row: dict[str, Any]) -> dict[str, Any]:
    return {str(key).strip().lower(): value for key, value in row.items()}


def first_value(row: dict[str, Any], *keys: str) -> Any:
    normalized = normalize_keyed_row(row)
    for key in keys:
        value = normalized.get(key.lower())
        if value is not None and str(value).strip() != "":
            return value
    return ""


def load_companies() -> list[dict[str, str]]:
    rows = sheet_rows("Companies")
    if not rows and workbook_state()["usingFallback"]:
        return SAMPLE_COMPANIES
    companies = []
    for row in rows:
        company = {
            "company_id": str(first_value(row, "company_id", "Company_ID")).strip(),
            "company_name": str(first_value(row, "company_name", "Company_Name", "name")).strip(),
            "short_name": str(first_value(row, "short_name", "Short_Name")).strip(),
            "status": str(first_value(row, "status", "Status") or "active").strip().lower(),
        }
        if company["company_id"] and company["company_name"]:
            companies.append(company)
    return companies


def active_companies() -> list[dict[str, str]]:
    return [company for company in load_companies() if company["status"] == "active"]


def load_report_recipients() -> list[dict[str, Any]]:
    rows = sheet_rows("Report_Recipients")
    if not rows:
        rows = sheet_rows("Report Recipients")
    if not rows and workbook_state()["usingFallback"]:
        return SAMPLE_REPORT_RECIPIENTS
    recipients = []
    for row in rows:
        item = {
            "company_id": str(first_value(row, "company_id", "Company_ID")).strip(),
            "report_type": str(first_value(row, "report_type", "Report_Type")).strip().lower(),
            "investor_id": str(first_value(row, "investor_id", "Investor_ID")).strip(),
            "active": truthy(first_value(row, "active", "Active")),
        }
        if item["company_id"] and item["report_type"] in report_type_ids() and item["investor_id"]:
            recipients.append(item)
    return recipients


def load_report_types() -> list[dict[str, str]]:
    rows = sheet_rows("Report_Types")
    if not rows:
        return [
            {"report_type": item, "label_en": REPORT_LABELS[item]["en"], "label_zh": REPORT_LABELS[item]["zh"], "active": True}
            for item in REPORT_ORDER
        ]
    report_types = []
    for row in rows:
        report_type = str(first_value(row, "report_type", "Report_Type", "type")).strip().lower()
        if not report_type:
            continue
        active_value = first_value(row, "active", "Active", "status", "Status")
        active = True if active_value == "" else truthy(active_value)
        if not active:
            continue
        fallback = REPORT_LABELS.get(report_type, {"en": report_type.title(), "zh": report_type})
        report_types.append(
            {
                "report_type": report_type,
                "label_en": str(first_value(row, "label_en", "name_en", "report_name_en", "English") or fallback["en"]).strip(),
                "label_zh": str(first_value(row, "label_zh", "name_zh", "report_name_zh", "Chinese") or fallback["zh"]).strip(),
                "active": True,
            }
        )
    return report_types or [
        {"report_type": item, "label_en": REPORT_LABELS[item]["en"], "label_zh": REPORT_LABELS[item]["zh"], "active": True}
        for item in REPORT_ORDER
    ]


def report_type_ids() -> list[str]:
    return [item["report_type"] for item in load_report_types()]


def report_type_label(report_type: str, language: str) -> str:
    for item in load_report_types():
        if item["report_type"] == report_type:
            return item["label_zh"] if language == "zh" else item["label_en"]
    fallback = REPORT_LABELS.get(report_type, {"en": report_type.title(), "zh": report_type})
    return fallback["zh"] if language == "zh" else fallback["en"]


def load_demo_batch() -> list[dict[str, Any]]:
    rows = sheet_rows("Demo_Batch")
    selected = []
    company_map = {company["company_id"]: company for company in load_companies()}
    for row in rows:
        if not truthy(first_value(row, "selected", "Selected")):
            continue
        company_id = str(first_value(row, "company_id", "Company_ID")).strip()
        report_type = str(first_value(row, "report_type", "Report_Type")).strip().lower()
        if not company_id or not report_type:
            continue
        selected.append(
            {
                "companyId": company_id,
                "companyName": company_map.get(company_id, {}).get("company_name", company_id),
                "reportType": report_type,
                "period": str(first_value(row, "period", "report_period", "Report_Period")).strip(),
                "selected": True,
                "pdfFileName": str(first_value(row, "pdf_file_name", "PDF_File_Name", "file_name", "File_Name")).strip(),
            }
        )
    return selected


def split_attachments(value: Any) -> list[str]:
    text = str(value or "").strip()
    if not text:
        return []
    return [
        item.strip().lstrip("-").strip()
        for item in re.split(r"[\n;,|]+", text)
        if item.strip().lstrip("-").strip()
    ]


def load_expected_email_preview() -> list[dict[str, Any]]:
    rows = sheet_rows("Expected_Email_Preview")
    expected = []
    for row in rows:
        investor_id = str(first_value(row, "investor_id", "Investor_ID")).strip()
        email = str(first_value(row, "email", "Email")).strip()
        investor_name = str(first_value(row, "investor_name", "Investor_Name")).strip()
        attachments = split_attachments(
            first_value(
                row,
                "expected_attachments",
                "attachment_file_names",
                "attachments",
                "attachment_list",
                "source_reports",
            )
        )
        if investor_id or email or investor_name or attachments:
            expected.append(
                {
                    "investorId": investor_id,
                    "investorName": investor_name,
                    "email": email,
                    "attachments": attachments,
                }
            )
    return expected


def compare_expected_preview(previews: list[dict[str, Any]]) -> dict[str, Any]:
    expected = load_expected_email_preview()
    actual_by_key = {
        (item.get("investorId") or item.get("email") or item.get("investorName")): item
        for item in previews
    }
    rows = []
    all_match = len(expected) == len(previews)
    for item in expected:
        key = item.get("investorId") or item.get("email") or item.get("investorName")
        actual = actual_by_key.get(key, {})
        expected_attachments = item.get("attachments", [])
        actual_attachments = actual.get("attachments", [])
        actual_report_keys = [
            f"{source['companyName']} {source['reportType']}"
            for source in actual.get("sourceReports", [])
        ]
        attachments_match = (
            not expected_attachments
            or sorted(expected_attachments) == sorted(actual_attachments)
            or sorted(expected_attachments) == sorted([source["label"] for source in actual.get("sourceReports", [])])
            or sorted([value.lower() for value in expected_attachments]) == sorted([value.lower() for value in actual_report_keys])
        )
        if not actual or not attachments_match:
            all_match = False
        rows.append(
            {
                "investorId": item.get("investorId", ""),
                "investorName": item.get("investorName", ""),
                "email": item.get("email", ""),
                "expectedAttachments": expected_attachments,
                "actualAttachments": actual_attachments,
                "actualSourceReports": [source["label"] for source in actual.get("sourceReports", [])],
                "actualReportKeys": actual_report_keys,
                "match": bool(actual) and attachments_match,
            }
        )
    return {
        "expectedInvestorCount": len(expected),
        "actualPreviewInvestorCount": len(previews),
        "matchesExpected": all_match,
        "rows": rows,
    }


def database_summary() -> dict[str, Any]:
    state = workbook_state()
    investors = load_investors()
    companies = active_companies()
    recipients = load_report_recipients()
    demo_batch = load_demo_batch()
    return {
        "workbookPath": state["path"],
        "usingFallback": state["usingFallback"],
        "fallbackWarning": state["warning"],
        "workbookError": state["error"],
        "missingRequiredSheets": state["missingRequiredSheets"],
        "sheetNames": state["sheetNames"],
        "companiesCount": len(companies),
        "activeInvestorsCount": len([item for item in investors if item["status"] == "active"]),
        "inactiveInvestorsCount": len([item for item in investors if item["status"] != "active"]),
        "reportRecipientsCount": len(recipients),
        "demoBatchSelectedReportsCount": len(demo_batch),
        "lastUploadedAt": LAST_UPLOADED_AT,
    }


def enriched_report_recipients() -> list[dict[str, Any]]:
    companies = {company["company_id"]: company for company in load_companies()}
    investors = {investor["investor_id"]: investor for investor in load_investors()}
    rows = []
    for item in load_report_recipients():
        company = companies.get(item["company_id"], {})
        investor = investors.get(item["investor_id"], {})
        rows.append(
            {
                "company_id": item["company_id"],
                "company_name": company.get("company_name", ""),
                "report_type": item["report_type"],
                "report_type_zh": report_type_label(item["report_type"], "zh"),
                "investor_id": item["investor_id"],
                "investor_name": investor.get("investor_name", ""),
                "email": investor.get("email", ""),
                "active": item["active"],
                "investor_status": investor.get("status", ""),
            }
        )
    return rows


def database_tables() -> dict[str, Any]:
    return {
        "companies": load_companies(),
        "investors": load_investors(),
        "reportRecipients": enriched_report_recipients(),
        "demoBatch": load_demo_batch(),
        "expectedEmailPreview": load_expected_email_preview(),
        "reportTypes": load_report_types(),
    }


def resolve_language(investor: dict[str, str], language_mode: str) -> str:
    if language_mode in {"zh", "en"}:
        return language_mode
    return "en" if investor.get("language") == "en" else "zh"


def report_label(company_name: str, report_type: str, language: str) -> str:
    return f"{company_name} {report_type_label(report_type, language)}"


def build_subject(fund_name: str, report_period: str, reports: list[dict[str, str]], language: str) -> str:
    if language == "zh":
        suffix = report_label(reports[0]["companyName"], reports[0]["reportType"], "zh") if len(reports) == 1 else "投資人報告"
        return f"{fund_name} {report_period} {suffix}"
    suffix = report_label(reports[0]["companyName"], reports[0]["reportType"], "en") if len(reports) == 1 else "Investor Reports"
    return f"{fund_name} {report_period} {suffix}"


def attachment_list(reports: list[dict[str, str]], language: str) -> str:
    return "\n".join(f"- {report_label(item['companyName'], item['reportType'], language)}" for item in reports)


def build_body(
    investor_name: str,
    fund_name: str,
    report_period: str,
    reports: list[dict[str, str]],
    language: str,
) -> str:
    attachments = attachment_list(reports, language)
    if language == "zh":
        return (
            f"親愛的 {investor_name} 您好，\n\n"
            f"隨信附上 {fund_name} 之 {report_period} 投資人報告，敬請參閱。\n\n"
            f"本次附件包含：\n{attachments}\n\n"
            "若有任何問題，請隨時與我們聯繫。\n\n"
            "敬祝 順心"
        )
    return (
        f"Dear {investor_name},\n\n"
        f"Please find attached the {report_period} investor report(s) for {fund_name}.\n\n"
        f"The attached files include:\n{attachments}\n\n"
        "Please let us know if you have any questions.\n\n"
        "Best regards,"
    )


def build_preview(
    investor: dict[str, str],
    fund_name: str,
    report_period: str,
    reports: list[dict[str, str]],
    language_mode: str,
) -> dict[str, Any]:
    language = resolve_language(investor, language_mode)
    attachment_file_names = [item["fileName"] for item in reports]
    return {
        "include": True,
        "investorId": investor["investor_id"],
        "investorName": investor["investor_name"],
        "email": investor["email"],
        "cc": investor["cc"],
        "subject": build_subject(fund_name, report_period, reports, language),
        "body": build_body(investor["investor_name"], fund_name, report_period, reports, language),
        "attachments": attachment_file_names,
        "attachmentCount": len(attachment_file_names),
        "sourceReports": [
            {
                "companyId": item["companyId"],
                "companyName": item["companyName"],
                "reportType": item["reportType"],
                "label": report_label(item["companyName"], item["reportType"], language),
            }
            for item in reports
        ],
        "language": language,
        "status": "ready" if investor["email"] else "blocked",
        "errorMessage": "" if investor["email"] else "Investor email is blank",
    }


def validate_payload(payload: dict[str, Any]) -> dict[str, Any]:
    fund_name = str(payload.get("fundName") or "").strip()
    report_period = str(payload.get("reportPeriod") or "").strip()
    selected_reports = payload.get("selectedReports") or []
    excluded_investor_ids = payload.get("excludedInvestorIds") or []
    language_mode = payload.get("languageMode") or "investor"

    if not fund_name:
        raise ValueError("基金／專案名稱不可空白。")
    if not report_period:
        raise ValueError("報告期間不可空白。")
    if not selected_reports:
        raise ValueError("請至少選擇一份本次要寄出的報告。")
    for item in selected_reports:
        if not item.get("companyId"):
            raise ValueError("選取的報告缺少公司 ID。")
        if item.get("reportType") not in report_type_ids():
            raise ValueError("選取的報告類型不支援。")
    if language_mode not in {"investor", "zh", "en"}:
        raise ValueError("不支援的語言模式。")

    return {
        "fundName": fund_name,
        "reportPeriod": report_period,
        "selectedReports": selected_reports,
        "excludedInvestorIds": excluded_investor_ids,
        "languageMode": language_mode,
    }


def report_file_field(company_id: str, report_type: str) -> str:
    safe_company_id = re.sub(r"[^A-Za-z0-9_-]", "_", company_id)
    return f"file_{safe_company_id}_{report_type}"


def selected_reports_with_files(payload: dict[str, Any], files: dict[str, UploadedFile]) -> list[SelectedReport]:
    company_map = {company["company_id"]: company for company in active_companies()}
    selected: list[SelectedReport] = []
    for item in payload["selectedReports"]:
        company_id = str(item["companyId"])
        report_type = str(item["reportType"])
        company = company_map.get(company_id)
        if not company:
            raise ValueError(f"公司不存在或未啟用：{company_id}")
        field_name = report_file_field(company_id, report_type)
        uploaded = files.get(field_name)
        if not uploaded:
            raise ValueError(f"請上傳 {company['company_name']} {report_type_label(report_type, 'zh')} 的 PDF。")
        if not uploaded.filename.lower().endswith(".pdf"):
            raise ValueError(f"{uploaded.filename} 不是有效的 PDF 檔案。")
        selected.append(
            SelectedReport(
                company_id=company_id,
                company_name=company["company_name"],
                report_type=report_type,
                file=uploaded,
            )
        )
    return selected


def lookup_summary_and_previews(
    payload: dict[str, Any],
    selected_reports: list[SelectedReport],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], dict[str, list[UploadedFile]]]:
    investors = active_investors()
    investor_map = {investor["investor_id"]: investor for investor in investors}
    mappings = [item for item in load_report_recipients() if item["active"]]
    excluded = set(payload.get("excludedInvestorIds") or [])
    grouped_reports: dict[str, list[dict[str, str]]] = {}
    grouped_files: dict[str, list[UploadedFile]] = {}
    summary = []

    for selected in selected_reports:
        rows = [
            item
            for item in mappings
            if item["company_id"] == selected.company_id
            and item["report_type"] == selected.report_type
            and item["investor_id"] in investor_map
        ]
        active_rows = [item for item in rows if item["investor_id"] not in excluded]
        status = "正常" if active_rows else "注意：尚未設定啟用中的固定收件人"
        summary.append(
            {
                "companyId": selected.company_id,
                "companyName": selected.company_name,
                "reportType": selected.report_type,
                "reportLabel": f"{selected.company_name} {report_type_label(selected.report_type, 'en')}",
                "recipientCount": len(active_rows),
                "pdfUploaded": True,
                "status": status,
            }
        )

        for row in active_rows:
            investor_id = row["investor_id"]
            grouped_reports.setdefault(investor_id, []).append(
                {
                    "companyId": selected.company_id,
                    "companyName": selected.company_name,
                    "reportType": selected.report_type,
                    "fileName": selected.file.filename,
                }
            )
            grouped_files.setdefault(investor_id, []).append(selected.file)

    previews = [
        build_preview(
            investor_map[investor_id],
            payload["fundName"],
            payload["reportPeriod"],
            reports,
            payload["languageMode"],
        )
        for investor_id, reports in grouped_reports.items()
    ]
    previews.sort(key=lambda item: item["investorName"].lower())
    return summary, previews, grouped_files


def build_raw_mime(
    to_address: str,
    cc: str,
    subject: str,
    body: str,
    attachments: list[UploadedFile],
) -> str:
    message = EmailMessage()
    message["To"] = to_address
    if cc:
        message["Cc"] = cc
    message["Subject"] = subject
    message.set_content(body, charset="utf-8")

    for attachment in attachments:
        maintype, subtype = (attachment.content_type or "application/pdf").split("/", 1)
        message.add_attachment(
            attachment.content,
            maintype=maintype,
            subtype=subtype,
            filename=attachment.filename,
        )

    return base64.urlsafe_b64encode(message.as_bytes()).decode("ascii").rstrip("=")


def gmail_access_token() -> str:
    required = ["GMAIL_CLIENT_ID", "GMAIL_CLIENT_SECRET", "GMAIL_REFRESH_TOKEN"]
    missing = [key for key in required if not os.getenv(key)]
    if missing:
        raise RuntimeError(f"Gmail API 環境變數尚未設定：{', '.join(missing)}")

    body = urllib.parse.urlencode(
        {
            "client_id": os.environ["GMAIL_CLIENT_ID"],
            "client_secret": os.environ["GMAIL_CLIENT_SECRET"],
            "refresh_token": os.environ["GMAIL_REFRESH_TOKEN"],
            "grant_type": "refresh_token",
        }
    ).encode()
    request_obj = urllib.request.Request(
        "https://oauth2.googleapis.com/token",
        data=body,
        headers={"Content-Type": "application/x-www-form-urlencoded"},
        method="POST",
    )
    with urllib.request.urlopen(request_obj, context=ssl.create_default_context(), timeout=30) as response:
        token_response = json.loads(response.read().decode("utf-8"))
    return token_response["access_token"]


def create_gmail_draft(raw_message: str) -> str:
    token = gmail_access_token()
    user_id = urllib.parse.quote(os.getenv("GMAIL_USER") or "me", safe="")
    request_obj = urllib.request.Request(
        f"https://gmail.googleapis.com/gmail/v1/users/{user_id}/drafts",
        data=json.dumps({"message": {"raw": raw_message}}).encode("utf-8"),
        headers={
            "Authorization": f"Bearer {token}",
            "Content-Type": "application/json",
        },
        method="POST",
    )
    with urllib.request.urlopen(request_obj, context=ssl.create_default_context(), timeout=45) as response:
        result = json.loads(response.read().decode("utf-8"))
    return result.get("id", "")


def append_send_log(entries: list[dict[str, Any]]) -> None:
    DATA_DIR.mkdir(exist_ok=True)
    with (DATA_DIR / "send-log.jsonl").open("a", encoding="utf-8") as file:
        for entry in entries:
            file.write(json.dumps(entry, ensure_ascii=False) + "\n")


def save_uploaded_workbook(uploaded: UploadedFile) -> dict[str, Any]:
    global CURRENT_WORKBOOK_PATH, LAST_UPLOADED_AT

    if not uploaded.filename.lower().endswith(".xlsx"):
        raise ValueError("只接受 .xlsx 格式的資料庫 Excel。")

    DATA_DIR.mkdir(exist_ok=True)
    target = DATA_DIR / "current_database.xlsx"
    target.write_bytes(uploaded.content)
    CURRENT_WORKBOOK_PATH = target
    LAST_UPLOADED_AT = datetime.now(timezone.utc).isoformat()
    (DATA_DIR / "current_database.meta.json").write_text(
        json.dumps({"path": str(target), "uploaded_at": LAST_UPLOADED_AT}, ensure_ascii=False),
        encoding="utf-8",
    )
    return {
        "summary": database_summary(),
        "tables": database_tables(),
    }


def templates_path() -> Path:
    return DATA_DIR / "email_templates.json"


def load_email_templates() -> dict[str, dict[str, str]]:
    path = templates_path()
    if not path.exists():
        return DEFAULT_TEMPLATES
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
        return {
            "zh": {**DEFAULT_TEMPLATES["zh"], **data.get("zh", {})},
            "en": {**DEFAULT_TEMPLATES["en"], **data.get("en", {})},
        }
    except Exception:
        return DEFAULT_TEMPLATES


def save_email_templates(templates: dict[str, Any]) -> dict[str, dict[str, str]]:
    DATA_DIR.mkdir(exist_ok=True)
    normalized = {
        "zh": {**DEFAULT_TEMPLATES["zh"], **templates.get("zh", {})},
        "en": {**DEFAULT_TEMPLATES["en"], **templates.get("en", {})},
    }
    templates_path().write_text(json.dumps(normalized, ensure_ascii=False, indent=2), encoding="utf-8")
    return normalized


def render_template(template: str, variables: dict[str, Any]) -> str:
    output = template
    for key, value in variables.items():
        output = output.replace("{{" + key + "}}", str(value or ""))
    return output


def simple_language(recipient: dict[str, str], language_mode: str) -> str:
    if language_mode == "zh":
        return "zh"
    if language_mode == "en":
        return "en"
    return "en" if recipient.get("language") == "en" else "zh"


def simple_attachment_list(report_types: list[str], language: str) -> str:
    return "\n".join(f"- {SIMPLE_REPORTS[item][language]}" for item in report_types)


def simple_report_names(report_types: list[str], language: str) -> str:
    return ", ".join(SIMPLE_REPORTS[item][language] for item in report_types)


def simple_file_ref(report_type: str) -> str:
    return str(SIMPLE_REPORTS[report_type]["path"])


def build_simple_previews(payload: dict[str, Any]) -> list[dict[str, Any]]:
    fund_name = str(payload.get("fundName") or "").strip()
    report_period = str(payload.get("reportPeriod") or "").strip()
    language_mode = str(payload.get("languageMode") or "recipient")
    recipients = payload.get("recipients") or []
    selection_matrix = payload.get("selectionMatrix") or {}
    templates = load_email_templates()

    if not fund_name:
        raise ValueError("基金／專案名稱不可空白。")
    if not report_period:
        raise ValueError("報告期間不可空白。")

    previews = []
    for recipient in recipients:
        recipient_id = recipient.get("recipient_id")
        selected_reports = [
            report_type
            for report_type in REPORT_ORDER
            if selection_matrix.get(recipient_id, {}).get(report_type)
        ]
        if not selected_reports:
            continue

        language = simple_language(recipient, language_mode)
        nickname = recipient.get("nickname") or recipient.get("name") or recipient.get("email")
        variables = {
            "fund_name": fund_name,
            "report_period": report_period,
            "recipient_name": recipient.get("name", ""),
            "nickname": nickname,
            "email": recipient.get("email", ""),
            "cc": recipient.get("cc", ""),
            "attachment_list": simple_attachment_list(selected_reports, language),
            "report_count": len(selected_reports),
            "report_names": simple_report_names(selected_reports, language),
        }
        template = templates[language]
        previews.append(
            {
                "include": bool(recipient.get("email")),
                "recipientId": recipient_id,
                "recipientName": recipient.get("name", ""),
                "nickname": recipient.get("nickname", ""),
                "email": recipient.get("email", ""),
                "cc": recipient.get("cc", ""),
                "language": language,
                "subject": render_template(template["subject"], variables),
                "body": render_template(template["body"], variables),
                "reportTypes": selected_reports,
                "attachments": [SIMPLE_REPORTS[item][language] for item in selected_reports],
                "attachmentFileRefs": [simple_file_ref(item) for item in selected_reports],
                "attachmentCount": len(selected_reports),
                "status": "ready" if recipient.get("email") else "blocked",
                "errorMessage": "" if recipient.get("email") else "缺少 Email",
            }
        )
    return previews


def uploaded_file_from_path(path_text: str, report_type: str) -> UploadedFile:
    path = Path(path_text)
    if not path.exists():
        raise ValueError(f"{SIMPLE_REPORTS[report_type]['zh']} PDF 尚未上傳。")
    return UploadedFile(
        field_name=f"file_{report_type}",
        filename=path.name,
        content_type="application/pdf",
        content=path.read_bytes(),
    )


def append_simple_send_log(batch_id: str, created_at: str, payload: dict[str, Any], results: list[dict[str, Any]]) -> None:
    entries = []
    for result in results:
        entries.append(
            {
                "batch_id": batch_id,
                "created_at": created_at,
                "fund_name": payload.get("fundName", ""),
                "report_period": payload.get("reportPeriod", ""),
                "recipient_id": result.get("recipientId", ""),
                "recipient_name": result.get("recipientName", ""),
                "email": result.get("email", ""),
                "selected_reports": result.get("reportTypes", []),
                "attachment_file_names": result.get("attachments", []),
                "gmail_draft_id": result.get("gmailDraftId", ""),
                "status": result.get("createStatus", ""),
                "error_message": result.get("errorMessage", ""),
            }
        )
    append_send_log(entries)


# ==============================================================================
# Flask app
# ==============================================================================

# Load env and runtime state at module import time (works for both gunicorn and direct execution)
load_env_file()
load_runtime_workbook_metadata()

APP_PASSWORD = os.getenv("APP_PASSWORD", "")

app = Flask(__name__)

# Derive a stable secret key:
# - Use SECRET_KEY env var if set (recommended for production)
# - Fall back to a hash of APP_PASSWORD so sessions survive gunicorn worker restarts
# - Fall back to random (sessions lost on restart; fine when no auth required)
_env_secret = os.getenv("SECRET_KEY")
if _env_secret:
    app.secret_key = _env_secret
elif APP_PASSWORD:
    app.secret_key = hashlib.sha256(f"investor-draft-tool-{APP_PASSWORD}".encode()).hexdigest()
else:
    app.secret_key = secrets.token_hex(32)

app.permanent_session_lifetime = timedelta(days=7)
app.config["MAX_CONTENT_LENGTH"] = 50 * 1024 * 1024  # 50 MB

# Ensure required directories exist (Railway container starts fresh each deploy)
for _d in [DATA_DIR, UPLOADS_DIR, UPLOADS_DIR / "recipients", UPLOADS_DIR / "reports", ROOT / "logs"]:
    _d.mkdir(parents=True, exist_ok=True)

if not APP_PASSWORD:
    print(
        "WARNING: APP_PASSWORD is not set. The app is running without password protection.",
        file=sys.stderr,
    )


# ------------------------------------------------------------------------------
# Auth middleware
# ------------------------------------------------------------------------------

def _requires_auth(path: str) -> bool:
    if not APP_PASSWORD:
        return False
    if path in ("/api/auth/login", "/login"):
        return False
    # Always allow static assets so the login page can load its styles/scripts
    if any(path.endswith(ext) for ext in (".css", ".js", ".ico", ".png", ".jpg", ".svg", ".woff", ".woff2")):
        return False
    return True


@app.before_request
def check_auth():
    if not _requires_auth(request.path):
        return
    if session.get("authenticated"):
        return
    if request.path.startswith("/api/"):
        return jsonify({"error": "請先登入", "requiresAuth": True}), 401
    return redirect("/login")


@app.route("/login")
def login_page():
    if not APP_PASSWORD or session.get("authenticated"):
        return redirect("/")
    return send_from_directory(STATIC_DIR, "login.html")


@app.route("/api/auth/login", methods=["POST"])
def auth_login():
    data = request.get_json() or {}
    if not APP_PASSWORD or data.get("password") == APP_PASSWORD:
        session["authenticated"] = True
        session.permanent = True
        return jsonify({"ok": True})
    return jsonify({"error": "密碼錯誤，請重試。"}), 401


@app.route("/api/auth/logout", methods=["POST"])
def auth_logout():
    session.clear()
    return jsonify({"ok": True})


# ------------------------------------------------------------------------------
# API routes
# ------------------------------------------------------------------------------

@app.route("/api/bootstrap")
def api_bootstrap():
    state = workbook_state()
    return jsonify({
        "investors": active_investors(),
        "companies": active_companies(),
        "reportRecipients": load_report_recipients(),
        "reportTypes": load_report_types(),
        "demoBatch": load_demo_batch(),
        "expectedEmailPreview": load_expected_email_preview(),
        "workbookPath": state["path"],
        "usingFallback": state["usingFallback"],
        "fallbackWarning": state["warning"],
        "workbookError": state["error"],
        "missingRequiredSheets": state["missingRequiredSheets"],
        "sheetNames": state["sheetNames"],
        "databaseSummary": database_summary(),
    })


@app.route("/api/database-summary")
def api_database_summary():
    return jsonify(database_summary())


@app.route("/api/database-tables")
def api_database_tables():
    return jsonify(database_tables())


@app.route("/api/templates", methods=["GET"])
def api_get_templates():
    return jsonify({"templates": load_email_templates()})


@app.route("/api/templates", methods=["POST"])
def api_save_templates():
    try:
        payload = request.get_json() or {}
        return jsonify({"templates": save_email_templates(payload.get("templates") or payload)})
    except Exception as exc:
        return jsonify({"error": str(exc)}), 400


@app.route("/api/upload-recipients", methods=["POST"])
def api_upload_recipients():
    try:
        if "recipients" not in request.files:
            return jsonify({"error": "請選擇要上傳的收件人 Excel。"}), 400
        f = request.files["recipients"]
        if not f.filename.lower().endswith(".xlsx"):
            return jsonify({"error": "只接受 .xlsx 格式的收件人資料表。"}), 400
        target_dir = UPLOADS_DIR / "recipients"
        target_dir.mkdir(parents=True, exist_ok=True)
        target = target_dir / "current_recipients.xlsx"
        f.save(str(target))
        recipients = parse_recipients_workbook(target)
        return jsonify({
            "recipients": recipients,
            "summary": recipients_summary(recipients),
            "fileRef": str(target),
        })
    except Exception as exc:
        return jsonify({"error": str(exc)}), 400


@app.route("/api/upload-report", methods=["POST"])
def api_upload_report():
    try:
        report_type = request.form.get("reportType", "")
        if report_type not in SIMPLE_REPORTS:
            return jsonify({"error": "不支援的報告類型。"}), 400
        if "report" not in request.files:
            return jsonify({"error": "請選擇要上傳的 PDF。"}), 400
        f = request.files["report"]
        if not f.filename.lower().endswith(".pdf"):
            return jsonify({"error": "只接受 .pdf 格式的報告檔案。"}), 400
        target = SIMPLE_REPORTS[report_type]["path"]
        target.parent.mkdir(parents=True, exist_ok=True)
        f.save(str(target))
        return jsonify({
            "reportType": report_type,
            "fileRef": str(target),
            "fileName": f.filename,
            "storedFileName": target.name,
        })
    except Exception as exc:
        return jsonify({"error": str(exc)}), 400


@app.route("/api/preview-simple", methods=["POST"])
def api_preview_simple():
    try:
        payload = request.get_json() or {}
        return jsonify({"previews": build_simple_previews(payload)})
    except Exception as exc:
        return jsonify({"error": str(exc)}), 400


@app.route("/api/create-drafts-simple", methods=["POST"])
def api_create_drafts_simple():
    batch_id = str(uuid.uuid4())
    created_at = datetime.now(timezone.utc).isoformat()
    try:
        payload = request.get_json() or {}
        previews = payload.get("previews") or build_simple_previews(payload)
        included = [p for p in previews if p.get("include") and p.get("status") == "ready"]
        results = []
        for preview in included:
            try:
                attachments = [
                    uploaded_file_from_path(file_ref, report_type)
                    for file_ref, report_type in zip(
                        preview.get("attachmentFileRefs", []),
                        preview.get("reportTypes", []),
                    )
                ]
                raw = build_raw_mime(
                    preview["email"],
                    preview.get("cc", ""),
                    preview["subject"],
                    preview["body"],
                    attachments,
                )
                draft_id = create_gmail_draft(raw)
                results.append({**preview, "createStatus": "success", "gmailDraftId": draft_id})
            except Exception as exc:
                results.append({**preview, "createStatus": "failed", "gmailDraftId": "", "errorMessage": str(exc)})
        append_simple_send_log(batch_id, created_at, payload, results)
        return jsonify({
            "batchId": batch_id,
            "successCount": sum(1 for r in results if r["createStatus"] == "success"),
            "failedCount": sum(1 for r in results if r["createStatus"] == "failed"),
            "results": results,
        })
    except Exception as exc:
        return jsonify({"error": str(exc)}), 400


@app.route("/api/upload-workbook", methods=["POST"])
def api_upload_workbook():
    try:
        if "workbook" not in request.files:
            return jsonify({"error": "請選擇要上傳的 .xlsx 資料庫 Excel。"}), 400
        f = request.files["workbook"]
        uploaded = UploadedFile(
            field_name="workbook",
            filename=f.filename,
            content_type=f.content_type or "application/octet-stream",
            content=f.read(),
        )
        return jsonify(save_uploaded_workbook(uploaded))
    except Exception as exc:
        return jsonify({"error": str(exc)}), 400


@app.route("/api/query-recipients-by-report", methods=["POST"])
def api_query_recipients_by_report():
    try:
        payload = request.get_json() or {}
        company_id = str(payload.get("companyId") or "").strip()
        report_type = str(payload.get("reportType") or "").strip().lower()
        rows = [
            row for row in enriched_report_recipients()
            if row["company_id"] == company_id and row["report_type"] == report_type and row["active"]
        ]
        return jsonify({"recipients": rows})
    except Exception as exc:
        return jsonify({"error": str(exc)}), 400


@app.route("/api/query-reports-by-investor", methods=["POST"])
def api_query_reports_by_investor():
    try:
        payload = request.get_json() or {}
        investor_id = str(payload.get("investorId") or "").strip()
        rows = [
            row for row in enriched_report_recipients()
            if row["investor_id"] == investor_id and row["active"]
        ]
        return jsonify({"reports": rows})
    except Exception as exc:
        return jsonify({"error": str(exc)}), 400


@app.route("/api/preview", methods=["POST"])
def api_preview():
    try:
        payload = validate_payload(json.loads(request.form.get("payload", "{}")))
        files = {
            key: UploadedFile(
                field_name=key,
                filename=f.filename,
                content_type=f.content_type or "application/pdf",
                content=f.read(),
            )
            for key, f in request.files.items()
        }
        selected_reports = selected_reports_with_files(payload, files)
        summary, previews, _ = lookup_summary_and_previews(payload, selected_reports)
        return jsonify({
            "lookupSummary": summary,
            "previews": previews,
            "expectedComparison": compare_expected_preview(previews),
        })
    except Exception as exc:
        return jsonify({"error": str(exc)}), 400


@app.route("/api/create-drafts", methods=["POST"])
def api_create_drafts():
    batch_id = str(uuid.uuid4())
    created_at = datetime.now(timezone.utc).isoformat()
    try:
        payload = validate_payload(json.loads(request.form.get("payload", "{}")))
        files = {
            key: UploadedFile(
                field_name=key,
                filename=f.filename,
                content_type=f.content_type or "application/pdf",
                content=f.read(),
            )
            for key, f in request.files.items()
        }
        selected_reports = selected_reports_with_files(payload, files)
        summary, previews, grouped_files = lookup_summary_and_previews(payload, selected_reports)
        results = []
        for preview in previews:
            if not preview["email"]:
                results.append({**preview, "createStatus": "failed", "gmailDraftId": "", "errorMessage": "Investor email is blank"})
                continue
            try:
                raw = build_raw_mime(
                    preview["email"],
                    preview["cc"],
                    preview["subject"],
                    preview["body"],
                    grouped_files[preview["investorId"]],
                )
                draft_id = create_gmail_draft(raw)
                results.append({**preview, "createStatus": "success", "gmailDraftId": draft_id})
            except Exception as exc:
                results.append({**preview, "createStatus": "failed", "gmailDraftId": "", "errorMessage": str(exc)})
        append_send_log([
            {
                "batch_id": batch_id,
                "created_at": created_at,
                "fund_name": payload["fundName"],
                "report_period": payload["reportPeriod"],
                "investor_id": result["investorId"],
                "investor_name": result["investorName"],
                "email": result["email"],
                "selected_reports": result["sourceReports"],
                "attachment_file_names": result["attachments"],
                "gmail_draft_id": result.get("gmailDraftId", ""),
                "status": result["createStatus"],
                "error_message": result.get("errorMessage", ""),
            }
            for result in results
        ])
        return jsonify({
            "batchId": batch_id,
            "successCount": sum(1 for r in results if r["createStatus"] == "success"),
            "failedCount": sum(1 for r in results if r["createStatus"] == "failed"),
            "results": results,
        })
    except Exception as exc:
        return jsonify({"error": str(exc)}), 400


# ------------------------------------------------------------------------------
# Error handlers
# ------------------------------------------------------------------------------

@app.errorhandler(413)
def request_entity_too_large(error):
    return jsonify({"error": "上傳檔案太大，單次上傳上限為 50MB。"}), 413


# ------------------------------------------------------------------------------
# Static file serving
# ------------------------------------------------------------------------------

@app.route("/")
def index():
    return send_from_directory(STATIC_DIR, "index.html")


@app.route("/<path:filename>")
def serve_static(filename):
    return send_from_directory(STATIC_DIR, filename)


# ------------------------------------------------------------------------------
# Entry point (local development)
# ------------------------------------------------------------------------------

def main() -> None:
    port = int(os.getenv("PORT", "3000"))
    print(f"投資人報告草稿產生器已啟動：http://127.0.0.1:{port}", file=sys.stderr)
    app.run(host="0.0.0.0", port=port, debug=False)


if __name__ == "__main__":
    main()
